# -*- coding: utf-8 -*-
"""LEARNING_CURRICULUM.md の内容を Excel に出力するスクリプト"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).resolve().parent.parent / "LEARNING_CURRICULUM.xlsx"

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header_row(ws, row: int, col_count: int):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def write_table(ws, start_row: int, headers: list, rows: list, col_widths: list | None = None) -> int:
    for i, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=i, value=h)
    style_header_row(ws, start_row, len(headers))
    r = start_row + 1
    for row_data in rows:
        for i, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=i, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
        r += 1
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return r


def sheet_overview(wb: Workbook):
    ws = wb.active
    ws.title = "概要"
    ws["A1"] = "10日間 実践学習カリキュラム"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:D1")

    info = [
        ("対象", "Cursor・Python・JavaScript・Next.js・MySQL・GitHub 初心者"),
        ("進め方", "小さな Web アプリを作りながら、毎日 2〜4 時間を目安に学習"),
        ("成果物", "非公開の「個人メモ＆タスク管理アプリ」（Reading & Task Board）"),
        ("公開", "しない（GitHub は Private リポジトリのみ）"),
    ]
    row = 3
    for label, value in info:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="作るアプリの機能").font = SUBHEADER_FONT
    row += 1
    write_table(
        ws,
        row,
        ["機能", "技術"],
        [
            ("メモ・タスクの一覧・追加・編集・削除", "Next.js（画面）+ Python API"),
            ("データ保存", "MySQL"),
            ("画面の動き・入力", "JavaScript / TypeScript"),
            ("開発環境", "Cursor"),
            ("履歴管理", "GitHub（private）"),
        ],
        [45, 35],
    )

    row = ws.max_row + 2
    ws.cell(row=row, column=1, value="最終構成").font = SUBHEADER_FONT
    row += 1
    arch = (
        "[ブラウザ] → Next.js (localhost:3000)\n"
        "        ↓ fetch\n"
        "    Python FastAPI (localhost:8000)\n"
        "        ↓\n"
        "    MySQL (localhost:3306)"
    )
    ws.cell(row=row, column=1, value=arch)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 80

    row = ws.max_row + 2
    ws.cell(row=row, column=1, value="推奨フォルダ構成（Day 10）").font = SUBHEADER_FONT
    row += 1
    folder = (
        "サンプルプロジェクト１/\n"
        "├── LEARNING_CURRICULUM.md / .xlsx\n"
        "├── REFLECTION.md\n"
        "├── README.md\n"
        "├── .gitignore\n"
        "├── frontend/  ← Next.js\n"
        "└── backend/   ← Python FastAPI"
    )
    ws.cell(row=row, column=1, value=folder)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 55


def sheet_daily_summary(wb: Workbook):
    ws = wb.create_sheet("10日間一覧")
    headers = ["日", "テーマ", "ゴール", "主な技術"]
    rows = [
        ("Day 0", "事前準備", "開発環境・Cursor・GitHub アカウント", "Cursor, Git, Node, Python, MySQL"),
        ("Day 1", "環境・GitHub・画面の箱", "Next.js でトップページ表示、Private リポジトリ", "Next.js, Git, GitHub, Cursor"),
        ("Day 2", "JS/TS + コンポーネント", "静的なタスク一覧 UI（ダミーデータ）", "JavaScript, TypeScript, React"),
        ("Day 3", "Python + FastAPI", "API が動き JSON を返す", "Python, FastAPI, HTTP"),
        ("Day 4", "fetch で API 接続", "Next.js から API 呼び出し", "fetch, async/await, 環境変数"),
        ("Day 5", "MySQL 永続化", "DB 保存、再起動後もデータ残る", "MySQL, SQL, Python DB"),
        ("Day 6", "CRUD 完成", "編集・削除・完了フラグ", "REST, SQL UPDATE/DELETE"),
        ("Day 7", "読書メモ機能", "books / book_notes テーブル", "DB設計, JOIN, 動的ルート"),
        ("Day 8", "Next.js 整理", "hooks・コンポーネント分割", "App Router, カスタムフック"),
        ("Day 9", "GitHub 実践", "ブランチ・PR・Issue、検索機能", "Git branch, PR, Issue"),
        ("Day 10", "仕上げ・振り返り", "完成・REFLECTION.md", "Docker概念, デプロイ概念"),
    ]
    write_table(ws, 1, headers, rows, [10, 28, 42, 28])
    ws.freeze_panes = "A2"


def sheet_day_detail(wb: Workbook):
    ws = wb.create_sheet("日別詳細")
    headers = ["日", "区分", "内容"]
    days = [
        (
            "Day 1",
            "ゴール",
            "リポジトリ作成、Next.js でトップページだけ表示",
            "学ぶこと",
            "ターミナル、フォルダ構成、npm / Git init・add・commit・push / Next.js app/page.tsx / Cursor でプロジェクト生成",
            "実践",
            "1) GitHub Private リポジトリ作成\n2) git init + create-next-app frontend\n3) npm run dev → localhost:3000\n4) page.tsx 編集\n5) README + push",
            "Cursor",
            "Chat: ヘッダー付きレイアウト / 1行ずつ import の意味を質問",
            "チェック",
            "□ localhost:3000 表示 □ GitHub Private に push □ 1回以上 commit",
        ),
        (
            "Day 2",
            "ゴール",
            "静的なタスク一覧 UI（まだ DB なし）",
            "学ぶこと",
            "const/let, map, オブジェクト / interface / useState / use client",
            "実践",
            "1) TaskList ダミーデータ 2) TaskForm + 追加（画面のみ） 3) Tailwind スタイル",
            "Cursor",
            "Inline Edit: 削除ボタン / useState コメント追記",
            "チェック",
            "□ map 表示 □ state に追加 □ commit: Day2 タスク一覧 UI",
        ),
        (
            "Day 3",
            "ゴール",
            "Python バックエンドが動き JSON を返す",
            "学ぶこと",
            "venv, pip, requirements.txt / FastAPI GET・POST / JSON / CORS",
            "実践",
            "1) backend + venv + fastapi uvicorn\n2) GET /health, GET/POST /tasks（メモリ）\n3) uvicorn --port 8000\n4) /docs で Swagger 確認",
            "Cursor",
            "CORS 最小コード / Traceback 貼って修正依頼",
            "チェック",
            "□ /docs で試せる □ POST で増える □ venv/ を gitignore",
        ),
        (
            "Day 4",
            "ゴール",
            "Next.js から Python API を呼び一覧・追加",
            "学ぶこと",
            "fetch, async/await, try/catch / NEXT_PUBLIC_API_URL / 別ポート開発",
            "実践",
            "1) .env.local に API URL\n2) useEffect で GET /tasks\n3) POST 後に再取得\n4) ローディング・エラー表示",
            "Cursor",
            "fetch エラーハンドリングコメント / DevTools Network",
            "チェック",
            "□ 両方起動で動作 □ リロードで消える（Day5で解決） □ commit Day4",
        ),
        (
            "Day 5",
            "ゴール",
            "タスクを DB に保存、再起動後も残る",
            "学ぶこと",
            "CREATE TABLE, INSERT, SELECT, UPDATE, DELETE / 主キー / mysql-connector-python",
            "実践",
            "1) learning_app DB + tasks テーブル\n2) backend/.env（gitignore）\n3) API を DB 操作に差し替え\n4) .env.example 作成",
            "Cursor",
            "（SQL 実行確認を AI に質問可）",
            "チェック",
            "□ SELECT * FROM tasks □ 再起動後も残る □ .env は commit しない",
        ),
        (
            "Day 6",
            "ゴール",
            "編集・削除・完了フラグまで一通り",
            "学ぶこと",
            "REST GET/POST/PUT/PATCH/DELETE / SQL WHERE, UPDATE / 編集UI",
            "実践",
            "API: PUT /tasks/{id}, PATCH done, DELETE\nフロント: 編集・削除・完了ボタン",
            "Cursor",
            "SQL インジェクション対策レビュー / パラメータ化クエリ",
            "チェック",
            "□ ブラウザから CRUD 全部 □ commit Day6",
        ),
        (
            "Day 7",
            "ゴール",
            "books テーブルとメモを追加",
            "学ぶこと",
            "1対多設計 / JOIN / 複数リソース API",
            "実践",
            "1) books, book_notes テーブル（CASCADE）\n2) /books, /books/{id}/notes\n3) app/books/[id]/page.tsx",
            "Cursor",
            "（動的ルートの説明を依頼可）",
            "チェック",
            "□ 本とメモ複数 □ CASCADE 削除 □ ER図を README に",
        ),
        (
            "Day 8",
            "ゴール",
            "コード整理、JavaScript 理解を深める",
            "学ぶこと",
            "layout.tsx, loading.tsx, 動的ルート / useTasks / 共通 UI",
            "実践",
            "1) hooks/useTasks.ts 2) lib/api.ts 3) ナビ 4) types/index.ts",
            "Cursor",
            "リファクタのみ・動作変更なし / diff 確認して Accept",
            "チェック",
            "□ README 構成と一致 □ commit Day8",
        ),
        (
            "Day 9",
            "ゴール",
            "ブランチ・PR・Issue を一人で体験",
            "学ぶこと",
            "git checkout -b / PR / Issue / バリデーション",
            "実践",
            "1) Issue 3つ 2) ブランチ→PR→Merge 3) 検索（LIKE or filter） 4) README セットアップ",
            "Cursor",
            "（README テンプレート生成依頼可）",
            "チェック",
            "□ PR merge 1回 □ Issue close □ README 完備",
        ),
        (
            "Day 10",
            "ゴール",
            "デモ完成、学習の言語化",
            "学ぶこと",
            "Docker Compose（任意）/ デプロイ概念 / 振り返り",
            "実践",
            "1) 空状態UI・スタイル統一・git tag v0.1.0（任意）\n2) REFLECTION.md\n3) 最終動作確認すべて",
            "Cursor",
            "振り返り文書の下書き依頼可",
            "チェック",
            "□ MySQL→Backend→Frontend □ タスクCRUD □ 本メモCRUD □ GitHub Private □ .env 未commit",
        ),
    ]
    ws.cell(row=1, column=1, value=headers[0])
    ws.cell(row=1, column=2, value=headers[1])
    ws.cell(row=1, column=3, value=headers[2])
    style_header_row(ws, 1, 3)

    row = 2
    for day_block in days:
        day_name = day_block[0]
        for i in range(1, len(day_block), 2):
            section = day_block[i]
            content = day_block[i + 1]
            ws.cell(row=row, column=1, value=day_name)
            ws.cell(row=row, column=2, value=section)
            ws.cell(row=row, column=3, value=content)
            for c in range(1, 4):
                cell = ws.cell(row=row, column=c)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = THIN_BORDER
                if section == "ゴール":
                    cell.fill = SUBHEADER_FILL
            row += 1
        row += 1  # blank between days
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A2"


def sheet_checklist(wb: Workbook):
    ws = wb.create_sheet("進捗チェックリスト")
    headers = ["日", "チェック項目", "完了"]
    items = [
        ("Day 0", "Cursor インストール", "□"),
        ("Day 0", "Node.js LTS インストール", "□"),
        ("Day 0", "Python 3.11+ インストール", "□"),
        ("Day 0", "MySQL または Docker+MySQL", "□"),
        ("Day 0", "Git インストール", "□"),
        ("Day 0", "GitHub アカウント作成", "□"),
        ("Day 0", "Cursor でプロジェクトフォルダを開いた", "□"),
        ("Day 1", "localhost:3000 でページ表示", "□"),
        ("Day 1", "GitHub Private に push", "□"),
        ("Day 1", "1回以上 commit", "□"),
        ("Day 2", "map で一覧表示", "□"),
        ("Day 2", "フォームから state に追加", "□"),
        ("Day 2", "commit: Day2 タスク一覧 UI", "□"),
        ("Day 3", "/docs で API 試行", "□"),
        ("Day 3", "POST でタスク追加（メモリ）", "□"),
        ("Day 3", "venv/ を gitignore", "□"),
        ("Day 4", "フロント+API 同時起動で動作", "□"),
        ("Day 4", "commit: Day4 連携", "□"),
        ("Day 5", "SELECT * FROM tasks 確認", "□"),
        ("Day 5", "再起動後もデータ残る", "□"),
        ("Day 5", ".env を commit していない", "□"),
        ("Day 6", "ブラウザから CRUD 全部", "□"),
        ("Day 6", "commit: Day6 CRUD 完成", "□"),
        ("Day 7", "本とメモを複数登録", "□"),
        ("Day 7", "CASCADE 削除確認", "□"),
        ("Day 7", "ER図を README に追記", "□"),
        ("Day 8", "フォルダ構成が README と一致", "□"),
        ("Day 8", "commit: Day8 整理", "□"),
        ("Day 9", "ブランチ→PR→merge 1回以上", "□"),
        ("Day 9", "Issue を close", "□"),
        ("Day 9", "README セットアップ手順完備", "□"),
        ("Day 10", "MySQL→Backend→Frontend 起動確認", "□"),
        ("Day 10", "タスク CRUD 最終確認", "□"),
        ("Day 10", "本・メモ CRUD 最終確認", "□"),
        ("Day 10", "REFLECTION.md 作成", "□"),
    ]
    write_table(ws, 1, headers, items, [10, 50, 8])
    ws.freeze_panes = "A2"
    note_row = ws.max_row + 2
    ws.cell(
        row=note_row,
        column=1,
        value="「完了」列の □ を ○ に書き換えるか、Excel のチェックボックスを挿入して進捗管理してください。",
    )
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=3)


def sheet_day0(wb: Workbook):
    ws = wb.create_sheet("Day0_事前準備")
    ws["A1"] = "事前準備（Day 0）"
    ws["A1"].font = Font(bold=True, size=14)

    prep = [
        "Cursor をインストール（日本語 UI 推奨）",
        "Node.js LTS をインストール（node -v / npm -v で確認）",
        "Python 3.11+ をインストール（python --version）",
        "MySQL または Docker Desktop + MySQL コンテナ",
        "Git をインストール（git --version）",
        "GitHub アカウント作成",
        "フォルダ「サンプルプロジェクト１」を Cursor で開く",
    ]
    row = 3
    for i, item in enumerate(prep, 1):
        ws.cell(row=row, column=1, value=f"□ {i}. {item}")
        row += 1

    row += 1
    write_table(
        ws,
        row,
        ["操作", "ショートカット（Win）"],
        [
            ("AI に質問・依頼", "Ctrl + L（Chat）"),
            ("コード選択して修正", "選択 → Ctrl + K（Inline Edit）"),
            ("ファイルを開く", "Ctrl + P"),
            ("ターミナル", "Ctrl + `"),
        ],
        [25, 30],
    )

    row = ws.max_row + 2
    ws.cell(row=row, column=1, value="AI への聞き方のコツ").font = SUBHEADER_FONT
    tips = [
        "「初心者向けにコメント付きで」「理由も説明して」と添える",
        "エラーは全文コピーして貼る",
        "1日の範囲は「このファイルだけ直して」と小さく区切る",
    ]
    row += 1
    for t in tips:
        ws.cell(row=row, column=1, value=f"・{t}")
        row += 1
    ws.column_dimensions["A"].width = 55


def sheet_cursor_github(wb: Workbook):
    ws = wb.create_sheet("GitHub・Git")
    write_table(
        ws,
        1,
        ["タイミング", "やること"],
        [
            ("Day 1", "Private リポジトリ作成、README だけ push"),
            ("Day 2〜9", "機能が動くたび git add → commit → push"),
            ("Day 10", "タグ v0.1.0 や Release（任意・非公開）"),
        ],
        [12, 50],
    )
    row = ws.max_row + 2
    ws.cell(row=row, column=1, value="毎日の Git ルーティン（5分）").font = SUBHEADER_FONT
    row += 1
    git_cmds = "git status\ngit add .\ngit commit -m \"Day3: タスク一覧 API を追加\"\ngit push"
    ws.cell(row=row, column=1, value=git_cmds)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    row += 2
    ws.cell(row=row, column=1, value="コミットメッセージ例").font = SUBHEADER_FONT
    row += 1
    for msg in [
        "Day1: プロジェクト初期構成",
        "Day5: MySQL に tasks テーブル作成",
        "fix: 作成日時が UTC になる問題を修正",
    ]:
        ws.cell(row=row, column=1, value=msg)
        row += 1
    ws.column_dimensions["A"].width = 55


def sheet_tech_map(wb: Workbook):
    ws = wb.create_sheet("技術マップ")
    write_table(
        ws,
        1,
        ["技術", "主に学ぶ日"],
        [
            ("Cursor", "毎日（特に Day 1, 8, 9）"),
            ("GitHub", "Day 1, 9, 10 + 毎日 commit"),
            ("JavaScript / TypeScript", "Day 2, 4, 8"),
            ("Next.js", "Day 1, 2, 4, 7, 8"),
            ("Python", "Day 3, 5, 6, 7"),
            ("MySQL", "Day 5, 6, 7"),
            ("HTTP / API 設計", "Day 3, 4, 6, 9"),
        ],
        [22, 40],
    )
    ws.freeze_panes = "A2"


def sheet_troubleshoot(wb: Workbook):
    ws = wb.create_sheet("つまずき対処")
    write_table(
        ws,
        1,
        ["症状", "確認すること"],
        [
            ("CORS エラー", "FastAPI の CORS 設定、NEXT_PUBLIC_API_URL"),
            ("DB 接続失敗", "MySQL 起動、.env のホスト/ユーザー/パスワード"),
            ("fetch failed", "backend がポート 8000 で動いているか"),
            ("ポート使用中", "他の node / uvicorn プロセスを終了"),
            ("Git push 拒否", "git remote -v、GitHub 認証（PAT または SSH）"),
        ],
        [18, 55],
    )
    row = ws.max_row + 2
    ws.cell(row=row, column=1, value="Cursor への質問例").font = SUBHEADER_FONT
    ws.cell(
        row=row + 1,
        column=1,
        value="「このエラーを初心者向けに原因と修正手順を3ステップで教えて: [エラー全文]」",
    )
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=2)


def sheet_next_steps(wb: Workbook):
    ws = wb.create_sheet("11日目以降")
    ws["A1"] = "学習後のおすすめステップ"
    ws["A1"].font = Font(bold=True, size=14)
    steps = [
        "認証: NextAuth または JWT（非公開のまま）",
        "テスト: Python pytest、フロント Vitest",
        "Docker Compose: MySQL + API を一発起動",
        "デプロイ体験: Vercel（フロント）+ Railway 等",
    ]
    for i, s in enumerate(steps, 3):
        ws.cell(row=i, column=1, value=f"{i-2}. {s}")

    row = len(steps) + 4
    ws.cell(row=row, column=1, value="毎日の開始ルーティン（2分）").font = SUBHEADER_FONT
    routine = [
        "Cursor でプロジェクトを開く",
        "MySQL → uvicorn → npm run dev の順で起動",
        "git pull（別PCなしならスキップ）",
        "今日の Day のチェックリストだけ読む",
    ]
    row += 1
    for r in routine:
        ws.cell(row=row, column=1, value=f"・{r}")
        row += 1
    ws.column_dimensions["A"].width = 60


def sheet_sql_reference(wb: Workbook):
    ws = wb.create_sheet("SQL参考")
    ws["A1"] = "Day 5: tasks テーブル"
    ws["A1"].font = SUBHEADER_FONT
    sql5 = (
        "CREATE DATABASE learning_app;\n"
        "USE learning_app;\n\n"
        "CREATE TABLE tasks (\n"
        "  id INT AUTO_INCREMENT PRIMARY KEY,\n"
        "  title VARCHAR(255) NOT NULL,\n"
        "  memo TEXT,\n"
        "  done TINYINT(1) DEFAULT 0,\n"
        "  created_at DATETIME DEFAULT CURRENT_TIMESTAMP\n"
        ");"
    )
    ws["A2"] = sql5
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 120

    ws["A4"] = "Day 7: books / book_notes"
    ws["A4"].font = SUBHEADER_FONT
    sql7 = (
        "CREATE TABLE books (\n"
        "  id INT AUTO_INCREMENT PRIMARY KEY,\n"
        "  title VARCHAR(255) NOT NULL,\n"
        "  author VARCHAR(255),\n"
        "  created_at DATETIME DEFAULT CURRENT_TIMESTAMP\n"
        ");\n\n"
        "CREATE TABLE book_notes (\n"
        "  id INT AUTO_INCREMENT PRIMARY KEY,\n"
        "  book_id INT NOT NULL,\n"
        "  content TEXT NOT NULL,\n"
        "  created_at DATETIME DEFAULT CURRENT_TIMESTAMP,\n"
        "  FOREIGN KEY (book_id) REFERENCES books(id) ON DELETE CASCADE\n"
        ");"
    )
    ws["A5"] = sql7
    ws["A5"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[5].height = 160
    ws.column_dimensions["A"].width = 80


def main():
    wb = Workbook()
    sheet_overview(wb)
    sheet_daily_summary(wb)
    sheet_day_detail(wb)
    sheet_checklist(wb)
    sheet_day0(wb)
    sheet_cursor_github(wb)
    sheet_tech_map(wb)
    sheet_troubleshoot(wb)
    sheet_next_steps(wb)
    sheet_sql_reference(wb)
    wb.save(OUTPUT)
    print(f"Created: {OUTPUT}")


if __name__ == "__main__":
    main()
